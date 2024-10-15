# pip install python-docx

from docx import Document
# import pandas as pd
import openpyxl as xl
import argparse
from pathlib import Path
import glob
import sys


def extract_tables_from_docx(file_path):
    """
    读取word 所有数据
    """
    try:
        # 打开 Word 文档
        doc = Document(file_path)

        # 初始化一个列表来存储所有表格的数据
        all_tables_data = []

        # 遍历文档中的所有表格
        for index, table in enumerate(doc.tables):
            # 初始化一个列表来存储当前表格的数据
            table_data = []
            # 遍历表格中的每一行
            for row in table.rows:
                # 初始化一个列表来存储当前行的数据
                row_data = []

                # 遍历行中的每一个单元格
                for cell in row.cells:
                    # 提取单元格中的文本并添加到行数据列表中
                    row_data.append(cell.text)

                # 将当前行的数据添加到表格数据列表中
                table_data.append(row_data)

            # 将当前表格的数据添加到所有表格数据列表中
            all_tables_data.append(table_data)
        return all_tables_data
    except Exception as e:
        print("\n读取文件：{},失败:{},请检查Word文件格式,必须是 docs".format(file_path, e))
        return None


def outputFile(wb: xl.Workbook, data: list, title: str):
    '''
    # 创建 DataFrame
    df = pd.DataFrame(data)
    # 将 DataFrame 写入 Excel 文件
    df.to_excel("example.xlsx", index=False)
    '''
    # 创建新的工作簿
    # wb =xl. Workbook()
    # ws = wb.active
    sheet = wb.create_sheet(title)
    # ws.title = title
    for row in data:
        # print("写单元格",*row)
        sheet.append(row)
    # wb.save("盘龙整改清单_程序自动生成.xlsx")


def progress_bar(i: float, title: str = ""):
    """
    进度条
    """
    i = int(i*100)
    print("\r", end="")
    print("{}: {}%: ".format(title, i))
    sys.stdout.flush()


def main(wb: xl.Workbook, file_path):

    # 提取表格数据
    tables_data = extract_tables_from_docx(file_path)
    if tables_data == None:
        return
    data = [["序号", "点位名称", "巡查日期", "巡查情况", "整改情况"]]
    # 打印提取的表格数据

    for i, table_data in enumerate(tables_data, start=1):
        # print(f"Table {i}:")
        progress_bar(i/len(tables_data), "处理：{}".format(file_path))

        name = None
        date = None
        info = None
        txt = None
        for row in table_data:
            if "点位名称" in row[0]:
                name = row[1]
            if "巡查日期" in row[0]:
                date = row[1]
            if "巡查情况" in row[0]:
                info = row[1]
            if "整改情况" in row[0]:
                txt = row[1]
            # print(row)
         
        data.append([i, name, date, info, txt])

        # print("\n")
    # print(*data)
    path = Path(file_path)
    outputFile(wb, data,  path.stem)


def find_files_with_extension(root_dir, extension):
    pattern = f"{root_dir}/**/*{extension}"
    found_files = glob.glob(pattern, recursive=True)
    return found_files


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="提取word表格中的文字,保存到Excell")
    parser.add_argument("-d", "--folder",  type=str, help="word文件所在目录")
    args = parser.parse_args()
    if args.folder == None:
        parser.print_help()
    else:
        words = find_files_with_extension(args. folder, ".docx")
        # 创建新的工作簿
        wb = xl. Workbook()
        for w in words:
            path = Path(w)
            if path.stem.startswith("~$"):
                continue
            print("处理文件：", w)

            main(wb, w)
        wb.save("整改清单_程序自动生成.xlsx")
